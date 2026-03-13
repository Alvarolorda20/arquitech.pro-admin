import json
import os
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

class GeneradorExcelComparativo:
    def __init__(self):
        # Estilos visuales (Restaurando el look & feel original)
        self.header_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # Gris Cabecera
        self.chapter_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid") # Gris Capitulo
        self.pauta_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")   # Verde Pauta
        self.oferta_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Amarillo Oferta
        self.alert_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")   # Rojo Alerta
        
        self.thin_border = Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        )
        self.bold_font = Font(bold=True)
        self.wrap_alignment = Alignment(wrap_text=True, vertical='center')
        self.center_alignment = Alignment(horizontal='center', vertical='center')
        self.right_alignment = Alignment(horizontal='right', vertical='center')
        
        # Formatos de número
        self.currency_fmt = '#,##0.00€'
        self.number_fmt = '#,##0.00'

    def _normalizar_id(self, codigo_completo):
        """Limpia IDs tipo '31::03.01' -> '03.01'."""
        if not codigo_completo:
            return ""
        return str(codigo_completo).split("::")[-1].strip()

    def cargar_datos(self, ruta_pauta, ruta_auditoria, ruta_oferta):
        """Carga datos y mapea totales de capítulos de la oferta."""
        
        # 1. Cargar Pauta
        with open(ruta_pauta, 'r', encoding='utf-8') as f:
            data_pauta = json.load(f)

        # 2. Cargar Oferta: Mapear Partidas Y Totales de Capítulo
        oferta_partidas_map = {}
        oferta_capitulos_map = {} # Nuevo: Para guardar los totales de capítulo
        
        with open(ruta_oferta, 'r', encoding='utf-8') as f:
            data_oferta = json.load(f)
            
        for cap in data_oferta:
            # Guardamos el total del capítulo indexado por su código (ej: "01", "02")
            cap_cod = str(cap.get('capitulo_codigo', '')).strip()
            if cap_cod:
                oferta_capitulos_map[cap_cod] = cap.get('total_capitulo', 0.0)

            # Indexamos las partidas
            for item in cap.get('partidas', []):
                cod_oferta = item.get('codigo')
                # Indexar por ID único compuesto (si existe) y simple
                if cap_cod and cod_oferta:
                    id_unico = f"{cap_cod}::{cod_oferta}"
                    oferta_partidas_map[id_unico] = item
                if cod_oferta:
                    oferta_partidas_map[cod_oferta] = item 

        # 3. Cargar Auditoría
        auditoria_map = {}
        with open(ruta_auditoria, 'r', encoding='utf-8') as f:
            data_audit = json.load(f)
            
        if isinstance(data_audit, list):
            for entry in data_audit:
                pauta_code = self._normalizar_id(entry.get('codigo_pauta'))
                auditoria_map[pauta_code] = entry
        
        return data_pauta, oferta_partidas_map, oferta_capitulos_map, auditoria_map

    def generar(self, ruta_pauta, ruta_mapping, ruta_auditoria, ruta_oferta, ruta_salida):
        print("Cargando y procesando datos...")
        pauta_data, oferta_map, oferta_caps_map, auditoria_map = self.cargar_datos(ruta_pauta, ruta_auditoria, ruta_oferta)
        
        wb = Workbook()
        ws = wb.active
        ws.title = "Comparativo Pauta-Oferta"

        # --- CABECERAS ---
        headers_sup = ["", "", "", "", "ESTIMACIÓN J&F (PAUTA)", "", "", "PROPUESTA PROVEEDOR", "", "", "", "AUDITORÍA"]
        ws.append(headers_sup)
        ws.merge_cells('E1:G1')
        ws.merge_cells('H1:K1')
        
        for cell in ws[1]:
            cell.fill = self.header_fill
            cell.font = self.bold_font
            cell.alignment = self.center_alignment

        headers_inf = [
            "Código", "Nat", "Ud", "Resumen Partida", 
            "CanPres", "PrPres", "ImpPres", 
            "Cód Oferta", "CanOferta", "PrOferta", "ImpOferta", 
            "Observaciones / Validación IA"
        ]
        ws.append(headers_inf)

        for cell in ws[2]:
            cell.border = self.thin_border
            cell.font = self.bold_font
            cell.alignment = self.center_alignment
            if cell.col_idx <= 4: cell.fill = self.header_fill
            elif 5 <= cell.col_idx <= 7: cell.fill = self.pauta_fill
            elif 8 <= cell.col_idx <= 11: cell.fill = self.oferta_fill
            else: cell.fill = self.header_fill

        # --- CUERPO ---
        current_row = 3
        
        for capitulo in pauta_data:
            cap_cod = str(capitulo.get('capitulo_codigo', '')).strip()
            
            # --- FILA DE CAPÍTULO (Restaurada) ---
            # 1. Código y Nombre
            ws.cell(row=current_row, column=1, value=cap_cod)
            ws.cell(row=current_row, column=4, value=capitulo.get('capitulo_nombre', '').upper())
            
            # 2. Totales de Capítulo (Pauta y Oferta)
            total_pauta = capitulo.get('total_capitulo', 0.0)
            total_oferta = oferta_caps_map.get(cap_cod, 0.0) # Recuperado del mapa de oferta
            
            c_pauta = ws.cell(row=current_row, column=7, value=total_pauta) # Total Pauta
            c_pauta.number_format = self.currency_fmt
            
            c_oferta = ws.cell(row=current_row, column=11, value=total_oferta) # Total Oferta
            c_oferta.number_format = self.currency_fmt

            # 3. Estilo de Fila de Capítulo
            for col in range(1, 13):
                c = ws.cell(row=current_row, column=col)
                c.fill = self.chapter_fill # Fondo Gris
                c.font = self.bold_font    # Negrita
                c.border = self.thin_border
                
            current_row += 1
            
            # --- PARTIDAS DEL CAPÍTULO ---
            for pauta_item in capitulo.get('partidas', []):
                code_pauta_raw = pauta_item.get('codigo_pauta', '')
                code_pauta_norm = self._normalizar_id(code_pauta_raw)
                
                audit_info = auditoria_map.get(code_pauta_norm, {})
                ids_oferta_asociados = audit_info.get('codigo_oferta', [])
                if isinstance(ids_oferta_asociados, str): ids_oferta_asociados = [ids_oferta_asociados]
                
                comentario = audit_info.get('comentario_tecnico', '') or audit_info.get('analisis', '')
                num_subfilas = max(1, len(ids_oferta_asociados))
                
                # Datos Pauta (Fila Principal)
                ws.cell(row=current_row, column=1, value=code_pauta_norm)
                ws.cell(row=current_row, column=2, value="Partida")
                ws.cell(row=current_row, column=3, value=pauta_item.get('unidad'))
                ws.cell(row=current_row, column=4, value=pauta_item.get('descripcion', '')[:300])
                
                ws.cell(row=current_row, column=5, value=pauta_item.get('cantidad')).number_format = self.number_fmt
                ws.cell(row=current_row, column=6, value=pauta_item.get('precio')).number_format = self.currency_fmt
                ws.cell(row=current_row, column=7, value=pauta_item.get('total')).number_format = self.currency_fmt
                
                # Subfilas para Ofertas (1 a N)
                for i in range(num_subfilas):
                    row_idx = current_row + i
                    if i > 0: ws.cell(row=row_idx, column=1, value=" ↳")
                    
                    if i < len(ids_oferta_asociados):
                        oferta_id_raw = ids_oferta_asociados[i]
                        item_oferta = oferta_map.get(oferta_id_raw) or oferta_map.get(self._normalizar_id(oferta_id_raw))

                        if item_oferta:
                            ws.cell(row=row_idx, column=8, value=item_oferta.get('codigo'))
                            ws.cell(row=row_idx, column=9, value=item_oferta.get('cantidad')).number_format = self.number_fmt
                            ws.cell(row=row_idx, column=10, value=item_oferta.get('precio')).number_format = self.currency_fmt
                            ws.cell(row=row_idx, column=11, value=item_oferta.get('total')).number_format = self.currency_fmt
                        else:
                            ws.cell(row=row_idx, column=8, value=f"{oferta_id_raw} (?)")
                    
                    # Observaciones
                    if i == 0:
                        cell_obs = ws.cell(row=row_idx, column=12, value=comentario)
                        if "❌" in comentario: cell_obs.fill = self.alert_fill

                # Bordes y alineación final de las filas creadas
                for r in range(current_row, current_row + num_subfilas):
                    for c in range(1, 13):
                        cell = ws.cell(row=r, column=c)
                        cell.border = self.thin_border
                        if c == 4 or c == 12: cell.alignment = self.wrap_alignment
                        elif c in [5,6,7,9,10,11]: cell.alignment = self.right_alignment # Números a la derecha
                        else: cell.alignment = self.center_alignment
                            
                current_row += num_subfilas

        # --- ANCHOS DE COLUMNA ---
        column_widths = {
            1: 10, 2: 8, 3: 5, 4: 50,      # Descripción
            5: 10, 6: 10, 7: 14,           # Importes Pauta
            8: 12, 9: 10, 10: 10, 11: 14,  # Importes Oferta
            12: 60                         # Observaciones
        }
        for col_idx, width in column_widths.items():
            ws.column_dimensions[get_column_letter(col_idx)].width = width

        # Guardar
        print(f"Guardando Excel en: {ruta_salida}")
        os.makedirs(os.path.dirname(ruta_salida), exist_ok=True)
        wb.save(ruta_salida)
        print("¡Excel generado correctamente!")

if __name__ == "__main__":
    # Prueba rápida
    gen = GeneradorExcelComparativo()
    gen.generar("mapped_pauta.json", "MAPPING_LINKS_FINAL_ricard.json", "AUDITORIA_VALIDADA_ricard.json", "FINAL_ricard.json", "test_comparativo.xlsx")